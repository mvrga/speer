import json
import re
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from pdf2image import convert_from_path
from PIL import Image
from pypdf import PdfReader
import pytesseract


@dataclass(frozen=True)
class EvidenceRecord:
    file_path: str
    file_name: str
    sha256: str
    evidence_type: str
    extraction_method: str
    text_preview: str
    invoice_number: str | None
    invoice_date: str | None
    total_amount: float | None
    currency: str | None
    iban: str | None
    bic: str | None
    status: str
    payment_ready: bool
    parse_errors: list[str]

    def to_dict(self) -> dict:
        return {
            "file_path": self.file_path,
            "file_name": self.file_name,
            "sha256": self.sha256,
            "evidence_type": self.evidence_type,
            "extraction_method": self.extraction_method,
            "text_preview": self.text_preview,
            "invoice_number": self.invoice_number,
            "invoice_date": self.invoice_date,
            "total_amount": self.total_amount,
            "currency": self.currency,
            "iban": self.iban,
            "bic": self.bic,
            "status": self.status,
            "payment_ready": self.payment_ready,
            "parse_errors": self.parse_errors,
        }


def _file_sha256(file_path: Path) -> str:
    hasher = sha256()
    with file_path.open("rb") as file_handle:
        for block in iter(lambda: file_handle.read(1024 * 1024), b""):
            hasher.update(block)
    return hasher.hexdigest()


def _extract_text_from_pdf(file_path: Path) -> str:
    reader = PdfReader(str(file_path))
    pages_text: list[str] = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        pages_text.append(page_text)
    return "\n".join(pages_text)


def _ocr_pdf(file_path: Path) -> str:
    images = convert_from_path(str(file_path))
    pages_text: list[str] = []
    for page_image in images:
        pages_text.append(pytesseract.image_to_string(page_image))
    return "\n".join(pages_text)


def _ocr_image(file_path: Path) -> str:
    image = Image.open(file_path)
    return pytesseract.image_to_string(image)


def _first_match(pattern: re.Pattern, text: str) -> str | None:
    match = pattern.search(text)
    if not match:
        return None
    return match.group(1).strip()


def _normalize_amount(raw_amount: str) -> float | None:
    cleaned = raw_amount.strip()
    cleaned = cleaned.replace("\u00a0", "").replace(" ", "")
    if "," in cleaned:
        cleaned = cleaned.replace(".", "")
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _amount_from_filename(file_name: str) -> float | None:
    amount_pattern = re.compile(
        r"([0-9]{1,3}(?:[._][0-9]{3})*(?:,[0-9]{2})|[0-9]+(?:,[0-9]{2})?)"
    )
    match = amount_pattern.search(file_name)
    if not match:
        return None
    return _normalize_amount(match.group(1))


def _clean_iban(raw_iban: str | None) -> str | None:
    if raw_iban is None:
        return None
    return re.sub(r"\s+", "", raw_iban)


def _iban_to_int_string(iban: str) -> str:
    converted = []
    for char in iban:
        if char.isdigit():
            converted.append(char)
        else:
            converted.append(str(ord(char) - 55))
    return "".join(converted)


def _is_valid_iban(iban: str) -> bool:
    if len(iban) < 15 or len(iban) > 34:
        return False
    if not iban[:2].isalpha():
        return False
    if not iban[2:].isalnum():
        return False
    rearranged = iban[4:] + iban[:4]
    numeric = _iban_to_int_string(rearranged)
    remainder = 0
    for digit in numeric:
        remainder = (remainder * 10 + int(digit)) % 97
    return remainder == 1


def _select_iban(text: str) -> str | None:
    iban_pattern = re.compile(r"\b([A-Z]{2}[0-9A-Z\s]{13,34})\b")
    for match in iban_pattern.findall(text.upper()):
        candidate = _clean_iban(match)
        if candidate and _is_valid_iban(candidate):
            return candidate
    return None


def _select_bic(text: str) -> str | None:
    bic_pattern = re.compile(r"\b([A-Z]{6}[A-Z0-9]{2}([A-Z0-9]{3})?)\b")
    banned = {"DESCRIPTION", "SECURITY"}
    for match in bic_pattern.findall(text.upper()):
        candidate = match[0]
        if candidate in banned:
            continue
        if len(candidate) not in {8, 11}:
            continue
        return candidate
    return None


def _safe_export_value(value: str | None, fallback: str) -> str:
    if value is None:
        return fallback
    cleaned = value.strip()
    return cleaned if cleaned else fallback


def _suggested_action(parse_errors: Iterable[str]) -> str:
    errors = set(parse_errors)
    actions: list[str] = []
    if "iban_missing" in errors or "iban_missing_or_invalid" in errors:
        actions.append("Fill IBAN")
    if "total_amount_missing" in errors or "total_amount_invalid" in errors:
        actions.append("Fill amount")
    if "invoice_number_missing" in errors:
        actions.append("Fill invoice number")
    if "invoice_date_missing" in errors:
        actions.append("Fill invoice date")
    if any(error.startswith("pdf_read_error") for error in errors):
        actions.append("Check PDF file")
    if any(error.startswith("ocr_error") for error in errors):
        actions.append("Review scan quality")
    if any(error.startswith("unsupported_format") for error in errors):
        actions.append("Convert to PDF")
    if any(error.startswith("non_pdf_evidence") for error in errors):
        actions.append("Provide PDF version")
    if not actions:
        actions.append("Review evidence manually")
    return ", ".join(actions)


def _parse_fields(text: str, file_path: Path) -> tuple[dict, list[str]]:
    parse_errors: list[str] = []

    invoice_number_pattern = re.compile(
        r"(?:invoice\s*number|invoice\s*no\.?|inv\.\s*no\.?|facture\s*no\.?|rechnungsnummer|rechnung[-\s]*nr\.?|belegnummer|vorgangsnummer)\s*[:#]?\s*([A-Z0-9\-/]+)",
        re.IGNORECASE,
    )
    invoice_date_pattern = re.compile(
        r"(?:invoice\s*date|date|rechnungsdatum)\s*[:#]?\s*([0-9]{4}[-/][0-9]{2}[-/][0-9]{2}|[0-9]{2}[-/][0-9]{2}[-/][0-9]{4}|[0-9]{2}\.[0-9]{2}\.[0-9]{4})",
        re.IGNORECASE,
    )
    total_amount_pattern = re.compile(
        r"(?:total\s*amount|amount\s*due|total|gesamtbetrag|rechnungsbetrag|zu\s*zahlen|summe)\s*[:#]?\s*([0-9][0-9\s.,]*)",
        re.IGNORECASE,
    )

    invoice_number = _first_match(invoice_number_pattern, text)
    invoice_date = _first_match(invoice_date_pattern, text)
    raw_amount = _first_match(total_amount_pattern, text)

    currency = "EUR"

    iban = _select_iban(text)
    if iban is None:
        parse_errors.append("iban_missing_or_invalid")

    bic = _select_bic(text)

    total_amount = None
    if raw_amount:
        total_amount = _normalize_amount(raw_amount)
        if total_amount is None:
            parse_errors.append("total_amount_invalid")
    else:
        total_amount = _amount_from_filename(file_path.name)
        if total_amount is None:
            parse_errors.append("total_amount_missing")
        else:
            parse_errors.append("total_amount_from_filename")

    if invoice_number is None:
        parse_errors.append("invoice_number_missing")

    if invoice_date is None:
        parse_errors.append("invoice_date_missing")

    if total_amount is not None and total_amount <= 0:
        parse_errors.append("total_amount_non_positive")

    return (
        {
            "invoice_number": invoice_number,
            "invoice_date": invoice_date,
            "total_amount": total_amount,
            "currency": currency,
            "iban": iban,
            "bic": bic,
        },
        parse_errors,
    )


def _build_record(
    file_path: Path,
    evidence_type: str,
    extraction_method: str,
    text: str,
    fields: dict,
    parse_errors: list[str],
) -> EvidenceRecord:
    total_amount = fields.get("total_amount")
    iban = fields.get("iban")
    currency = fields.get("currency")

    payment_ready = bool(total_amount and total_amount > 0 and iban and currency)
    status = "ok" if payment_ready and fields.get("invoice_number") else "needs_review"

    return EvidenceRecord(
        file_path=str(file_path),
        file_name=file_path.name,
        sha256=_file_sha256(file_path),
        evidence_type=evidence_type,
        extraction_method=extraction_method,
        text_preview=text[:1200],
        invoice_number=fields.get("invoice_number"),
        invoice_date=fields.get("invoice_date"),
        total_amount=total_amount,
        currency=currency,
        iban=iban,
        bic=fields.get("bic"),
        status=status,
        payment_ready=payment_ready,
        parse_errors=parse_errors,
    )


def parse_pdf(file_path: Path) -> EvidenceRecord:
    try:
        text = _extract_text_from_pdf(file_path)
    except Exception as exc:  # noqa: BLE001
        return _build_record(
            file_path,
            evidence_type="pdf_text",
            extraction_method="pdf_text",
            text="",
            fields={
                "invoice_number": None,
                "invoice_date": None,
                "total_amount": None,
                "currency": "EUR",
                "iban": None,
                "bic": None,
            },
            parse_errors=[f"pdf_read_error:{exc}"],
        )

    if len(text.strip()) < 40:
        try:
            ocr_text = _ocr_pdf(file_path)
        except Exception as exc:  # noqa: BLE001
            fields, errors = _parse_fields(text, file_path)
            errors.append(f"ocr_error:{exc}")
            return _build_record(
                file_path,
                evidence_type="pdf_scan",
                extraction_method="pdf_text",
                text=text,
                fields=fields,
                parse_errors=errors,
            )

        fields, errors = _parse_fields(ocr_text, file_path)
        return _build_record(
            file_path,
            evidence_type="pdf_scan",
            extraction_method="ocr",
            text=ocr_text,
            fields=fields,
            parse_errors=errors,
        )

    fields, errors = _parse_fields(text, file_path)
    return _build_record(
        file_path,
        evidence_type="pdf_text",
        extraction_method="pdf_text",
        text=text,
        fields=fields,
        parse_errors=errors,
    )


def parse_image(file_path: Path) -> EvidenceRecord:
    try:
        text = _ocr_image(file_path)
    except Exception as exc:  # noqa: BLE001
        return _build_record(
            file_path,
            evidence_type="image",
            extraction_method="ocr",
            text="",
            fields={
                "invoice_number": None,
                "invoice_date": None,
                "total_amount": None,
                "currency": "EUR",
                "iban": None,
                "bic": None,
            },
            parse_errors=[f"ocr_error:{exc}"],
        )

    fields, errors = _parse_fields(text, file_path)
    return _build_record(
        file_path,
        evidence_type="image",
        extraction_method="ocr",
        text=text,
        fields=fields,
        parse_errors=errors,
    )


def parse_xml(file_path: Path) -> EvidenceRecord:
    return _build_record(
        file_path,
        evidence_type="xml",
        extraction_method="xml",
        text="",
        fields={
            "invoice_number": None,
            "invoice_date": None,
            "total_amount": None,
            "currency": "EUR",
            "iban": None,
            "bic": None,
        },
        parse_errors=["xml_parsing_not_implemented"],
    )


def detect_and_parse(file_path: Path) -> EvidenceRecord:
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return parse_pdf(file_path)
    if suffix in {".png", ".jpg", ".jpeg"}:
        return parse_image(file_path)
    if suffix == ".xml":
        return parse_xml(file_path)

    return _build_record(
        file_path,
        evidence_type="unknown",
        extraction_method="unknown",
        text="",
        fields={
            "invoice_number": None,
            "invoice_date": None,
            "total_amount": None,
            "currency": "EUR",
            "iban": None,
            "bic": None,
        },
        parse_errors=[f"unsupported_format:{suffix or 'unknown'}"],
    )


def process_invoice(file_path: str | Path) -> dict:
    record = detect_and_parse(Path(file_path))
    return record.to_dict()


def parse_invoice_files(file_paths: Iterable[Path]) -> list[EvidenceRecord]:
    records: list[EvidenceRecord] = []
    for file_path in file_paths:
        records.append(detect_and_parse(file_path))
    return records


def export_xlsx(records: Iterable[EvidenceRecord], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "speer_evidence"

    headers = [
        "file_path",
        "file_name",
        "sha256",
        "evidence_type",
        "extraction_method",
        "invoice_number",
        "invoice_date",
        "total_amount",
        "currency",
        "iban",
        "bic",
        "status",
        "payment_ready",
        "parse_errors",
    ]
    worksheet.append(headers)

    for record in records:
        worksheet.append(
            [
                record.file_path,
                record.file_name,
                record.sha256,
                record.evidence_type,
                record.extraction_method,
                record.invoice_number,
                record.invoice_date,
                record.total_amount,
                record.currency,
                record.iban,
                record.bic,
                record.status,
                record.payment_ready,
                ";".join(record.parse_errors),
            ]
        )

    workbook.save(output_path)


def export_payment_instructions(records: Iterable[EvidenceRecord], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "payment_instructions"

    headers = [
        "beneficiary_name",
        "iban",
        "bic",
        "amount",
        "currency",
        "reference",
    ]
    worksheet.append(headers)

    for record in records:
        if not record.payment_ready:
            continue
        beneficiary_name = _safe_export_value(
            getattr(record, "beneficiary_name", None), "UNKNOWN"
        )
        reference = _safe_export_value(record.invoice_number, record.file_name)
        reference = _safe_export_value(reference, record.file_name)
        worksheet.append(
            [
                beneficiary_name,
                record.iban,
                record.bic,
                record.total_amount,
                record.currency,
                reference,
            ]
        )

    workbook.save(output_path)


def export_review_pack(records: Iterable[EvidenceRecord], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "review"

    headers = [
        "file_name",
        "status",
        "extraction_method",
        "total_amount",
        "currency",
        "iban",
        "invoice_number",
        "text_preview",
        "suggested_action",
        "parse_errors",
    ]
    worksheet.append(headers)

    for record in records:
        if record.status != "needs_review":
            continue
        worksheet.append(
            [
                record.file_name,
                record.status,
                record.extraction_method,
                record.total_amount,
                record.currency,
                record.iban,
                record.invoice_number,
                record.text_preview,
                _suggested_action(record.parse_errors),
                ";".join(record.parse_errors),
            ]
        )

    workbook.save(output_path)


def export_json_audit(records: Iterable[EvidenceRecord], output_path: Path, run_id: str) -> dict:
    timestamp = datetime.now(timezone.utc).isoformat()
    payload = {
        "run_id": run_id,
        "timestamp": timestamp,
        "records": [record.to_dict() for record in records],
    }

    with output_path.open("w", encoding="utf-8") as file_handle:
        json.dump(payload, file_handle, indent=2, ensure_ascii=False)

    return payload


def export_review_json(records: Iterable[EvidenceRecord], output_path: Path, run_id: str) -> dict:
    timestamp = datetime.now(timezone.utc).isoformat()
    payload = {
        "run_id": run_id,
        "timestamp": timestamp,
        "records": [
            record.to_dict() for record in records if record.status == "needs_review"
        ],
    }

    with output_path.open("w", encoding="utf-8") as file_handle:
        json.dump(payload, file_handle, indent=2, ensure_ascii=False)

    return payload


def export_outputs(
    records: Iterable[EvidenceRecord],
    xlsx_output_path: str | Path,
    json_output_path: str | Path,
    ok_xlsx_output_path: str | Path,
    review_xlsx_output_path: str | Path,
    review_json_output_path: str | Path,
    run_id: str | None = None,
) -> dict:
    resolved_run_id = run_id or str(uuid.uuid4())
    export_xlsx(records, Path(xlsx_output_path))
    export_payment_instructions(records, Path(ok_xlsx_output_path))
    export_review_pack(records, Path(review_xlsx_output_path))
    export_json_audit(records, Path(json_output_path), resolved_run_id)
    export_review_json(records, Path(review_json_output_path), resolved_run_id)

    return {
        "run_id": resolved_run_id,
        "xlsx": str(xlsx_output_path),
        "json": str(json_output_path),
        "ok_xlsx": str(ok_xlsx_output_path),
        "review_xlsx": str(review_xlsx_output_path),
        "review_json": str(review_json_output_path),
    }


def process_evidence_files(
    file_paths: Iterable[str | Path],
    xlsx_output_path: str | Path,
    json_output_path: str | Path,
    ok_xlsx_output_path: str | Path,
    review_xlsx_output_path: str | Path,
    review_json_output_path: str | Path,
    run_id: str | None = None,
) -> list[dict]:
    resolved_paths = [Path(path).expanduser().resolve() for path in file_paths]
    records = parse_invoice_files(resolved_paths)
    export_outputs(
        records,
        xlsx_output_path,
        json_output_path,
        ok_xlsx_output_path,
        review_xlsx_output_path,
        review_json_output_path,
        run_id,
    )
    return [record.to_dict() for record in records]


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Speer evidence parser")
    parser.add_argument("inputs", nargs="+", help="Evidence files")
    parser.add_argument("--xlsx", required=True, help="Output XLSX path")
    parser.add_argument("--json", required=True, help="Output JSON audit log path")
    parser.add_argument(
        "--ok-xlsx", required=True, help="Output XLSX path for payment-ready invoices"
    )
    parser.add_argument(
        "--review-xlsx",
        required=True,
        help="Output XLSX path for needs_review invoices",
    )
    parser.add_argument(
        "--review-json",
        required=True,
        help="Output JSON path for needs_review invoices",
    )
    parser.add_argument("--run-id", help="Run ID for audit log")
    args = parser.parse_args()

    process_evidence_files(
        args.inputs,
        args.xlsx,
        args.json,
        args.ok_xlsx,
        args.review_xlsx,
        args.review_json,
        args.run_id,
    )
